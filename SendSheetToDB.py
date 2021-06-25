import mysql.connector
from openpyxl import load_workbook
workbook = load_workbook(r"C:\Users\Amruta\HNeLibrary\DummyData_HNLibrary_Eng.xlsx")

sheet=workbook.active

conn=mysql.connector.connect(
            host='localhost',
            user='root',
            passwd='ma5t3rb1a5t3r',
            database='hnelibrary'
        )
conn.autocommit=True
cursor=conn.cursor(buffered=True)


query="INSERT INTO tblbookinfo (ID, AccessionNo, BookType, SubjectNo, Title, Author, Publisher, Price, Currency, Pages, Language, CupboardNo, RackNo, BarcodePrint) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"
rows=list(sheet.iter_rows(min_row=6,min_col=1,max_col=14, values_only=True))
for row in rows:
    
    cursor.execute(query, (row[0],row[1],row[2], row[3], row[4], row[5], row[6], row[7],row[8],row[9], row[10], row[11], row[12], row[13]))
    #print(row)

conn.close()



    
        
